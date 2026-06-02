"""Unit tests. Run with: pytest

Fully offline. Exercises the deterministic pieces (policy parsing, gap analysis,
crosswalk, eval scoring) and the mapper's guardrails against hallucinated ids,
using the framework-aware stub client. No API key or network required.
"""

from __future__ import annotations

import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))
sys.path.insert(0, str(ROOT))

from control_mapper.analysis import analyze  # noqa: E402
from control_mapper.crosswalk import (  # noqa: E402
    CrosswalkInput,
    build_crosswalk,
    equivalences_for,
)
from control_mapper.frameworks import load_framework, parse_policy  # noqa: E402
from control_mapper.mapper import ControlMapper  # noqa: E402
from eval.evaluate import load_gold, score  # noqa: E402
from eval.stub_client import StubClient  # noqa: E402

FW_DIR = ROOT / "data" / "frameworks"
NIST = load_framework(FW_DIR / "nist_800-53_subset.json")
ISO = load_framework(FW_DIR / "iso_27001_2022_annexa_subset.json")
SOC2 = load_framework(FW_DIR / "soc2_tsc_subset.json")
ALL_FRAMEWORKS = [NIST, ISO, SOC2]
POLICY = (ROOT / "data" / "sample_policies" / "sample_policy.md").read_text()


def test_parse_policy_splits_numbered_statements():
    statements = parse_policy(POLICY)
    assert len(statements) == 10
    assert statements[0].id == "P1"
    assert "\n" not in statements[1].text  # wrapped lines joined


def test_all_frameworks_load_with_expected_size():
    for fw in ALL_FRAMEWORKS:
        assert len(fw.controls) == 21
        # control ids are unique within a framework
        assert len(fw.control_ids()) == len(fw.controls)


def test_mapper_drops_hallucinated_control_ids():
    class BadClient:
        def get_mappings(self, system, tool, user_content):
            return [
                {"statement_id": "P1", "control_id": "AT-2", "coverage": "full",
                 "confidence": 0.9, "rationale": "valid"},
                {"statement_id": "P1", "control_id": "ZZ-99", "coverage": "full",
                 "confidence": 0.9, "rationale": "invented id"},
                {"statement_id": "P99", "control_id": "AT-2", "coverage": "full",
                 "confidence": 0.9, "rationale": "invented statement"},
            ]

    statements = parse_policy(POLICY)
    result = ControlMapper(BadClient()).map(statements, NIST)
    pairs = {(m.statement_id, m.control_id) for m in result.mappings}
    assert ("P1", "AT-2") in pairs
    assert ("P1", "ZZ-99") not in pairs
    assert ("P99", "AT-2") not in pairs


def test_stub_is_framework_aware():
    statements = parse_policy(POLICY)
    nist_pairs = {(m.statement_id, m.control_id)
                  for m in ControlMapper(StubClient()).map(statements, NIST).mappings}
    iso_pairs = {(m.statement_id, m.control_id)
                 for m in ControlMapper(StubClient()).map(statements, ISO).mappings}
    assert ("P1", "AT-2") in nist_pairs       # NIST control id
    assert ("P1", "A.6.3") in iso_pairs       # ISO control id
    assert ("P1", "AT-2") not in iso_pairs    # frameworks do not bleed into each other


def test_analysis_flags_unmapped_statement_for_nist():
    statements = parse_policy(POLICY)
    result = ControlMapper(StubClient()).map(statements, NIST)
    analysis = analyze(result, NIST, statements)
    # P9 (cyber insurance) has no NIST control in scope.
    assert "P9" in analysis.unmapped_statements
    assert 0.0 <= analysis.coverage_ratio <= 1.0
    assert set(analysis.covered_controls).isdisjoint(analysis.partial_controls)


def test_eval_detects_known_errors_across_frameworks():
    statements = parse_policy(POLICY)
    expectations = {
        "NIST_SP_800-53_R5": (("P5", "SC-7"), ("P2", "AC-3")),
        "ISO_IEC_27001_2022": (("P5", "A.8.20"), ("P2", "A.5.18")),
        "SOC2_TSC_2017": (("P6", "CC4.1"), ("P9", "CC9.1")),
    }
    for fw in ALL_FRAMEWORKS:
        result = ControlMapper(StubClient()).map(statements, fw)
        gold_pairs, gold_cov, _ = load_gold(fw.framework_id)
        s = score(result, gold_pairs, gold_cov)
        expected_fp, expected_fn = expectations[fw.framework_id]
        assert expected_fp in s.false_positives, fw.framework_id
        assert expected_fn in s.false_negatives, fw.framework_id
        assert 0.0 < s.precision <= 1.0
        assert 0.0 < s.recall < 1.0


def test_crosswalk_anchors_same_obligation_across_frameworks():
    statements = parse_policy(POLICY)
    inputs = [
        CrosswalkInput(fw, ControlMapper(StubClient()).map(statements, fw))
        for fw in ALL_FRAMEWORKS
    ]
    crosswalk = build_crosswalk(statements, inputs)
    p3 = next(r for r in crosswalk.rows if r.statement_id == "P3")
    # The "MFA for remote access" obligation maps to a control in every framework.
    for fw in ALL_FRAMEWORKS:
        assert p3.by_framework[fw.framework_id], fw.framework_id


def test_equivalences_link_controls_via_shared_statements():
    statements = parse_policy(POLICY)
    inputs = [
        CrosswalkInput(fw, ControlMapper(StubClient()).map(statements, fw))
        for fw in ALL_FRAMEWORKS
    ]
    crosswalk = build_crosswalk(statements, inputs)
    equivs = equivalences_for(crosswalk, "NIST_SP_800-53_R5")
    # NIST AC-17 (remote access) shares statement P3 with an ISO control.
    assert "AC-17" in equivs
    iso_targets = equivs["AC-17"]["ISO_IEC_27001_2022"]
    assert any("P3" in stmts for stmts in iso_targets.values())


def test_csv_export_round_trips(tmp_path):
    import csv

    from control_mapper.export import crosswalk_to_csv

    statements = parse_policy(POLICY)
    inputs = [
        CrosswalkInput(fw, ControlMapper(StubClient()).map(statements, fw))
        for fw in ALL_FRAMEWORKS
    ]
    crosswalk = build_crosswalk(statements, inputs)
    out = tmp_path / "cw.csv"
    n = crosswalk_to_csv(crosswalk, statements, inputs, out)
    assert n > 0
    with open(out, newline="", encoding="utf-8") as fh:
        rows = list(csv.DictReader(fh))
    assert len(rows) == n
    # Every framework appears, and every control id is real.
    seen = {r["framework_id"] for r in rows}
    assert seen == {fw.framework_id for fw in ALL_FRAMEWORKS}
    valid = set().union(*[fw.control_ids() for fw in ALL_FRAMEWORKS])
    assert all(r["control_id"] in valid for r in rows)


def test_xlsx_export_creates_expected_sheets(tmp_path):
    openpyxl = __import__("importlib").util.find_spec("openpyxl")
    if openpyxl is None:
        import pytest

        pytest.skip("openpyxl not installed")
    from openpyxl import load_workbook

    from control_mapper.export import crosswalk_to_xlsx

    statements = parse_policy(POLICY)
    inputs = [
        CrosswalkInput(fw, ControlMapper(StubClient()).map(statements, fw))
        for fw in ALL_FRAMEWORKS
    ]
    crosswalk = build_crosswalk(statements, inputs)
    out = tmp_path / "cw.xlsx"
    crosswalk_to_xlsx(crosswalk, statements, inputs, out)
    wb = load_workbook(out)
    assert set(wb.sheetnames) == {"Summary", "Mappings", "Crosswalk", "Gaps"}
