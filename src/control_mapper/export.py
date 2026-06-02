"""Export crosswalk results to CSV and formatted XLSX.

Two outputs, both aimed at a GRC reader who lives in Excel:

- CSV  : a flat, long-format table (one row per mapping) that drops straight into
         any GRC tool, BI tool, or pivot table.
- XLSX : a formatted workbook with four sheets -- Summary, Crosswalk (wide,
         one row per obligation), Mappings (long, auditable detail with rationale),
         and Gaps (controls no statement covered, per framework).

Design notes for reviewers:
- Detail (control title, coverage, confidence, rationale) comes from the original
  mappings, not just the crosswalk rows, so the export is fully auditable.
- The Summary sheet's headline counts are written as Excel formulas (COUNTIF /
  COUNTA over the Mappings sheet) rather than hardcoded, so the workbook stays
  live if a reviewer edits the detail. Run the xlsx skill's recalc to populate
  cached values.
"""

from __future__ import annotations

import csv
from dataclasses import dataclass
from pathlib import Path
from typing import List

from .crosswalk import Crosswalk, CrosswalkInput


@dataclass
class _Record:
    statement_id: str
    statement_text: str
    framework_id: str
    control_id: str
    control_title: str
    coverage: str
    confidence: float
    rationale: str


def _build_records(
    statements, inputs: List[CrosswalkInput]
) -> List[_Record]:
    """Flatten every framework's mappings into one long-format list of records."""
    text_by_id = {s.id: s.text for s in statements}
    records: List[_Record] = []
    for inp in inputs:
        fw = inp.framework
        for m in inp.result.mappings:
            ctrl = fw.get(m.control_id)
            records.append(
                _Record(
                    statement_id=m.statement_id,
                    statement_text=text_by_id.get(m.statement_id, ""),
                    framework_id=fw.framework_id,
                    control_id=m.control_id,
                    control_title=ctrl.title if ctrl else "",
                    coverage=m.coverage.value,
                    confidence=round(m.confidence, 2),
                    rationale=m.rationale,
                )
            )
    # Stable, human-friendly order: statement, then framework, then control.
    records.sort(key=lambda r: (_statement_sort_key(r.statement_id), r.framework_id, r.control_id))
    return records


def _statement_sort_key(statement_id: str):
    # "P10" should sort after "P9", so sort by the trailing integer when present.
    digits = "".join(ch for ch in statement_id if ch.isdigit())
    return (int(digits) if digits else 0, statement_id)


CSV_HEADER = [
    "statement_id",
    "statement_text",
    "framework_id",
    "control_id",
    "control_title",
    "coverage",
    "confidence",
    "rationale",
]


def crosswalk_to_csv(
    crosswalk: Crosswalk, statements, inputs: List[CrosswalkInput], path: str | Path
) -> int:
    """Write the long-format crosswalk to CSV. Returns the number of rows written."""
    records = _build_records(statements, inputs)
    with open(path, "w", newline="", encoding="utf-8") as fh:
        writer = csv.writer(fh)
        writer.writerow(CSV_HEADER)
        for r in records:
            writer.writerow(
                [
                    r.statement_id,
                    r.statement_text,
                    r.framework_id,
                    r.control_id,
                    r.control_title,
                    r.coverage,
                    r.confidence,
                    r.rationale,
                ]
            )
    return len(records)


# ---------------------------------------------------------------------------
# XLSX
# ---------------------------------------------------------------------------

_FONT_NAME = "Arial"


def crosswalk_to_xlsx(
    crosswalk: Crosswalk, statements, inputs: List[CrosswalkInput], path: str | Path
) -> None:
    """Write a formatted, multi-sheet workbook. Requires openpyxl."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    header_font = Font(name=_FONT_NAME, bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F3A5F")
    base_font = Font(name=_FONT_NAME)
    title_font = Font(name=_FONT_NAME, bold=True, size=14)
    wrap = Alignment(vertical="top", wrap_text=True)
    top = Alignment(vertical="top")
    full_fill = PatternFill("solid", fgColor="E2EFDA")     # soft green
    partial_fill = PatternFill("solid", fgColor="FFF2CC")  # soft amber
    gap_fill = PatternFill("solid", fgColor="FCE4E4")      # soft red

    records = _build_records(statements, inputs)
    framework_ids = crosswalk.framework_ids

    wb = Workbook()

    def style_header(ws, ncols: int) -> None:
        for c in range(1, ncols + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    # --- Mappings sheet (long, auditable) -- built first so Summary can reference it.
    ms = wb.active
    ms.title = "Mappings"
    ms.append(CSV_HEADER)
    for r in records:
        ms.append(
            [r.statement_id, r.statement_text, r.framework_id, r.control_id,
             r.control_title, r.coverage, r.confidence, r.rationale]
        )
    style_header(ms, len(CSV_HEADER))
    for row in ms.iter_rows(min_row=2):
        for cell in row:
            cell.font = base_font
            cell.alignment = wrap if cell.column in (2, 8) else top
        cov_cell = row[5]
        cov_cell.fill = full_fill if cov_cell.value == "full" else partial_fill
    widths = {1: 12, 2: 48, 3: 20, 4: 12, 5: 30, 6: 11, 7: 11, 8: 52}
    for col, w in widths.items():
        ms.column_dimensions[get_column_letter(col)].width = w
    ms.freeze_panes = "A2"
    ms.auto_filter.ref = f"A1:H{ms.max_row}"
    last = ms.max_row  # last data row in Mappings; used by Summary formulas

    # --- Crosswalk sheet (wide, one row per obligation) ---
    cs = wb.create_sheet("Crosswalk")
    header = ["Statement ID", "Statement"] + framework_ids
    cs.append(header)
    text_by_id = {s.id: s.text for s in statements}
    for row in crosswalk.rows:
        cells = [row.statement_id, text_by_id.get(row.statement_id, "")]
        for fid in framework_ids:
            pairs = row.by_framework.get(fid, [])
            cells.append(
                "; ".join(f"{c} [{cov}]" for c, cov in pairs)
                if pairs
                else "(gap / out of scope)"
            )
        cs.append(cells)
    style_header(cs, len(header))
    for row in cs.iter_rows(min_row=2):
        for cell in row:
            cell.font = base_font
            cell.alignment = wrap
        for cell in row[2:]:
            if cell.value == "(gap / out of scope)":
                cell.fill = gap_fill
    cs.column_dimensions["A"].width = 12
    cs.column_dimensions["B"].width = 50
    for i in range(len(framework_ids)):
        cs.column_dimensions[get_column_letter(3 + i)].width = 26
    cs.freeze_panes = "C2"

    # --- Gaps sheet (uncovered controls per framework) ---
    gs = wb.create_sheet("Gaps")
    gs.append(["Framework", "Control ID", "Control Title", "Status"])
    for inp in inputs:
        fw = inp.framework
        touched = {m.control_id for m in inp.result.mappings}
        for ctrl in fw.controls:
            if ctrl.id not in touched:
                gs.append([fw.framework_id, ctrl.id, ctrl.title, "uncovered"])
    style_header(gs, 4)
    for row in gs.iter_rows(min_row=2):
        for cell in row:
            cell.font = base_font
            cell.alignment = top
            cell.fill = gap_fill
    gs.column_dimensions["A"].width = 22
    gs.column_dimensions["B"].width = 12
    gs.column_dimensions["C"].width = 40
    gs.column_dimensions["D"].width = 12
    gs.freeze_panes = "A2"

    # --- Summary sheet (counts via formulas over Mappings) ---
    summ = wb.create_sheet("Summary", 0)  # first sheet
    summ["A1"] = "Policy-to-Control Crosswalk Summary"
    summ["A1"].font = title_font
    summ["A3"] = (
        "AI-assisted draft. A qualified human must review every mapping before use "
        "as audit evidence. Counts below are live Excel formulas over the Mappings sheet."
    )
    summ["A3"].font = Font(name=_FONT_NAME, italic=True)
    summ["A3"].alignment = wrap
    summ.merge_cells("A3:E3")

    head_row = 5
    headers = ["Framework", "Controls in framework", "Mappings", "Full", "Partial"]
    for c, h in enumerate(headers, start=1):
        cell = summ.cell(row=head_row, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    size_by_fw = {inp.framework.framework_id: len(inp.framework.controls) for inp in inputs}
    for i, fid in enumerate(framework_ids):
        r = head_row + 1 + i
        summ.cell(row=r, column=1, value=fid).font = base_font
        summ.cell(row=r, column=2, value=size_by_fw.get(fid, 0)).font = base_font
        # Mappings (total rows for this framework) via COUNTIF on Mappings!C
        summ.cell(row=r, column=3,
                  value=f'=COUNTIF(Mappings!$C$2:$C${last},A{r})').font = base_font
        # Full / Partial via COUNTIFS on framework + coverage
        summ.cell(row=r, column=4,
                  value=f'=COUNTIFS(Mappings!$C$2:$C${last},A{r},Mappings!$F$2:$F${last},"full")').font = base_font
        summ.cell(row=r, column=5,
                  value=f'=COUNTIFS(Mappings!$C$2:$C${last},A{r},Mappings!$F$2:$F${last},"partial")').font = base_font
    summ.column_dimensions["A"].width = 24
    for col in ("B", "C", "D", "E"):
        summ.column_dimensions[col].width = 20

    wb.save(str(path))
